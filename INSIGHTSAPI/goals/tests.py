"""This module contains the tests for the goals app."""

import openpyxl
from django.conf import settings
from django.contrib.auth.models import Permission
from django.core.files.uploadedfile import SimpleUploadedFile
from django.db.models import Q
from django.urls import reverse
from django.utils import timezone
from rest_framework import status

from services.tests import BaseTestCase
from users.models import User

from .models import Goals, TableInfo


# * Recuerda que cuando corres los tests, NO se envía un correo electrónico real.
class GoalAPITestCase(BaseTestCase):
    """Test the Goal API."""

    def setUp(self):
        """Set up the test client and the URL for the goal list view."""
        super().setUp()
        user = User.objects.get(username="staffnet")
        permission_view = Permission.objects.get(codename="view_goals")
        user.user_permissions.add(permission_view)
        permission_add = Permission.objects.get(codename="add_goals")
        # permission_update = Permission.objects.get(codename="change_goals")
        user.user_permissions.add(permission_add)
        user.user_permissions.add(permission_view)
        # user.user_permissions.add(permission_update)
        permission_view_history = Permission.objects.get(
            codename="view_historicalgoals"
        )
        user.user_permissions.add(permission_view_history)

    def test_serializer(self):
        """Test the Goal serializer."""
        self.test_claro_upload()
        response = self.client.get(reverse("goal-list"))
        self.assertEqual(response.status_code, 200)

    def test_get_my_goal_without_permission(self):
        """Test the get-my-goals view."""
        self.user.user_permissions.clear()
        # Create a goal object
        Goals.objects.create(
            cedula=self.user.cedula,
            name="Heibert",
            campaign_goal="Base Test Goal",
            result="50",
            evaluation="50",
            quality="50",
            clean_desk="50",
            total="50",
            job_title_goal="Developer",
            last_update=timezone.now(),
            criteria_goal="50",
            quantity_goal="50",
            goal_date="ENERO-2022",
            execution_date="FEBRERO-2022",
        )
        response = self.client.get(f"/goals/{self.user.cedula}/")
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data.get("cedula"), self.user.cedula)  # type: ignore

    def test_metas_upload(self, called=False):
        """Test the upload-excel view."""

        file_path = "./utils/excels/Entrega de metas-ENERO-2018.xlsx"
        # Add the user cedula to the file
        wb = openpyxl.load_workbook(filename=file_path)
        ws = wb.active
        if not ws:
            self.fail("No active sheet found")
        ws["B21"] = self.user.cedula
        wb.save(file_path)
        # Create a SimpleUploadedFile instance from the Excel file
        with open(file_path, "rb") as file_obj:
            file_data = file_obj.read()
        excel_file = SimpleUploadedFile(
            "Entrega de metas-ENERO-2018.xlsx",
            file_data,
            content_type="application/vnd.ms-excel",
        )
        # Send the POST request to the upload-excel URL with the Excel file data
        response = self.client.post(
            reverse("goal-list"), {"file": excel_file}, cookies=self.client.cookies  # type: ignore
        )
        # Assert the response status code and perform additional assertions for the response data
        number_goals = Goals.objects.all().count()
        # print(response.data)
        self.assertEqual(response.status_code, 201)
        self.assertTrue(number_goals > 0)

    def test_metas_upload_without_permission(self):
        """Test the upload-excel view."""
        self.user.user_permissions.clear()
        file_path = "utils/excels/Entrega de metas-ENERO-2018.xlsx"
        with open(file_path, "rb") as file_obj:
            file_data = file_obj.read()
        excel_file = SimpleUploadedFile(
            "Entrega de metas-ENERO-2018.xlsx",
            file_data,
            content_type="application/vnd.ms-excel",
        )
        response = self.client.post(
            reverse("goal-list"), {"file": excel_file}, cookies=self.client.cookies  # type: ignore
        )
        self.assertEqual(response.status_code, 403)

    def test_execution_upload(self, called=False):
        """Test the upload-excel view."""
        if called:
            # Create a SimpleUploadedFile instance from the Excel file
            file_path = "utils/excels/Ejecución de metas-enerO-2022.xlsx"
            # Add the user cedula to the file
            wb = openpyxl.load_workbook(filename=file_path, data_only=True)
            ws = wb.active
            if not ws:
                self.fail("No active sheet found")
            ws["B340"] = self.user.cedula
            wb.save(file_path)
            # Create a SimpleUploadedFile instance from the Excel file
            with open(file_path, "rb") as file_obj:
                file_data = file_obj.read()
            excel_file = SimpleUploadedFile(
                "Ejecución de metas-enerO-2022.xlsx",
                file_data,
                content_type="application/vnd.ms-excel",
            )
            # Send the POST request to the upload-excel URL with the Excel file data
            response = self.client.post(reverse("goal-list"), {"file": excel_file})
            # Assert the response status code and perform additional assertions for the response data
            self.assertEqual(response.status_code, 201)
            count = Goals.objects.exclude(total="").count()
            self.assertTrue(count > 0)

    def test_borrado_accepted_with_another_upload(self):
        """Test the upload-excel view."""
        delete = Permission.objects.get(codename="delete_goals")
        self.user.user_permissions.add(delete)
        # Invoke the test_metas_upload() method to have data in the database and some goals like accepted
        self.test_metas_upload(called=True)
        # See if there are goals created
        number_goals = Goals.objects.all().count()
        self.assertTrue(number_goals > 0)
        # put accepted to True and accepted_at to now() to Goals
        Goals.objects.all().update(accepted=True, accepted_at=timezone.now())
        # Read the file again and upload it
        file_path = "utils/excels/Entrega de metas-ENERO-2018.xlsx"
        with open(file_path, "rb") as file_obj:
            file_data = file_obj.read()
        excel_file = SimpleUploadedFile(
            "Entrega de metas-ENERO-2018.xlsx",
            file_data,
            content_type="application/vnd.ms-excel",
        )
        # Send the POST request to the upload-excel URL with the Excel file data
        response = self.client.post(reverse("goal-list"), {"file": excel_file})
        self.assertEqual(response.status_code, 201)
        # See if the accepted goals were deleted
        first_goal = Goals.objects.exclude(accepted_at=None).first()
        self.assertIsNone(first_goal)
        count = Goals.objects.exclude(Q(accepted__isnull=True) | Q(accepted="")).count()
        count_at = Goals.objects.exclude(accepted_at=None).count()
        self.assertEqual((count, count_at), (0, 0))
        # Do the same verifications but with execution
        Goals.objects.all().update(
            accepted_execution=True, accepted_execution_at=timezone.now()
        )
        self.test_execution_upload(called=True)
        count_execution = Goals.objects.exclude(
            Q(accepted__isnull=True) | Q(accepted="")
        ).count()
        first_goal = Goals.objects.exclude(accepted_execution_at=None).first()
        self.assertIsNone(first_goal, first_goal)
        count_at_execution = Goals.objects.exclude(accepted_execution_at=None).count()
        self.assertEqual((count_execution, count_at_execution), (0, 0))

    def test_claro_upload(self):
        """Test the upload-excel view."""
        # Create a SimpleUploadedFile instance from the Excel file
        file_path = "utils/excels/Entrega de metas ejemplo Claro-ENERO-2028.xlsx"
        with open(file_path, "rb") as file_obj:
            file_data = file_obj.read()
        excel_file = SimpleUploadedFile(
            "Entrega de metas ejemplo Claro-ENERO-2028.xlsx",
            file_data,
            content_type="application/vnd.ms-excel",
        )
        # Send the POST request to the upload-excel URL with the Excel file data
        response = self.client.post(reverse("goal-list"), {"file": excel_file})
        # Assert the response status code and perform additional assertions for the response data
        self.assertEqual(response.status_code, 201)
        self.assertTrue(Goals.objects.all().count() > 0)
        self.assertTrue(TableInfo.objects.all().count() > 0)
        self.assertTrue(Goals.objects.all().exclude(table_goal=None).count() > 0)

    def test_get_history(self):
        """Test the get-history view."""
        self.test_claro_upload()
        self.test_execution_upload(called=True)
        # Check with a month that has no goals
        response = self.client.get("/goals/?date=DICIembre-2028&column=delivery")
        self.assertEqual(response.status_code, 200, response.data)
        self.assertFalse(len(response.data) > 0)  # type: ignore
        # Check with a month that has goals
        response = self.client.get("/goals/?date=ENERO-2028&column=delivery")
        self.assertEqual(response.status_code, 200)
        self.assertEqual(len(response.data), 110)  # type: ignore
        self.assertIn("ENERO-2028", response.data[0].get("goal_date"))  # type: ignore
        # Check with a month that has execution
        response = self.client.get("/goals/?date=ENERO-2022&column=execution")
        self.assertEqual(response.status_code, 200)
        self.assertEqual(len(response.data), 339)  # type: ignore
        # Upload a file to test if the same month distinct between delivery and execution
        file_path = "utils/excels/Entrega de metas ejemplo Claro-ENERO-2028.xlsx"
        with open(file_path, "rb") as file_obj:
            file_data = file_obj.read()
        excel_file = SimpleUploadedFile(
            "Entrega de metas ejemplo Claro-ENERO-2022.xlsx",
            file_data,
            content_type="application/vnd.ms-excel",
        )
        # Send the POST request to the upload-excel URL with the Excel file data
        response = self.client.post(reverse("goal-list"), {"file": excel_file})
        self.assertEqual(response.status_code, 201)
        # Check with a month that has twice
        response = self.client.get("/goals/?date=ENERO-2022&column=delivery")
        self.assertEqual(response.status_code, 200)
        self.assertEqual(len(response.data), 110)  # type: ignore
        self.assertIn("ENERO-2022", response.data[0].get("goal_date"))  # type: ignore

    def test_get_only_my_history(self):
        """Test the get-history view."""
        self.test_metas_upload(called=True)
        self.user.user_permissions.clear()
        response = self.client.get("/goals/?date=ENERO-2022&column=delivery")
        self.assertEqual(response.status_code, 200)
        self.assertEqual(len(response.data), 1)  # type: ignore
        self.assertIn("STAFFNET PRUEBAS", response.data[0].get("name"))  # type: ignore

    def test_get_one_history(self):
        """Test the get-one-history view."""
        self.test_claro_upload()
        self.test_execution_upload(called=True)
        # Check with a month that has goals
        response = self.client.get("/goals/1016080155/?date=ENERO-2028&column=delivery")
        self.assertEqual(response.status_code, 200, response.data)
        self.assertGreater(len(response.data), 0)  # type: ignore
        self.assertEqual(response.data["cedula"], "1016080155", response.data)  # type: ignore
        self.assertIn("ENERO-2028", response.data.get("goal_date"))  # type: ignore
        self.assertEqual(len(response.data), 28)  # type: ignore
        # Check with a month that has execution
        response = self.client.get(
            "/goals/1192792468/?date=ENERO-2022&column=execution"
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data["cedula"], "1192792468")  # type: ignore
        self.assertEqual(len(response.data), 28)  # type: ignore
        self.assertIn("ENERO-2022", response.data.get("execution_date"))  # type: ignore

    def test_get_campaigns(self):
        """Test the get-campaigns view."""
        self.test_claro_upload()
        response = self.client.get("/goals/?campaign=CLARO")
        self.assertEqual(response.status_code, 200)
        self.assertEqual(len(response.data), 110)

    def test_patch_goal_delivery(self):
        """Test the update-goal view."""

        # Create a goal object
        Goals.objects.create(
            cedula=settings.TEST_CEDULA,
            name="Heibert",
            campaign_goal="Base Test Goal",
            result="50",
            evaluation="50",
            quality="50",
            clean_desk="50",
            total="50",
            job_title_goal="Developer",
            last_update=timezone.now(),
            criteria_goal="50",
            quantity_goal="50",
            goal_date="ENERO-2022",
            execution_date="FEBRERO-2022",
        )
        # Prepare the request data
        payload = {
            "accepted": True,
        }
        # Send a PATCH request to the update-goal view
        response = self.client.patch(
            "/goals/" + str(settings.TEST_CEDULA) + "/", data=payload
        )
        # Assert the response status code and content
        self.assertEqual(response.status_code, status.HTTP_200_OK, response.data)
        # Assert that the goal was updated with the new data
        self.assertEqual(
            response.data["message"], "La meta fue aceptada.", response.data
        )

    def test_patch_goal_delivery_claro(self):
        """Test the update-goal view."""
        # Get the first goal from the database
        table_info_data = [
            {
                "name": "ANTIGUOS",
                "fringe": "CERO",
                "diary_goal": "70",
                "days": "24",
                "month_goal": "1680",
                "hours": "8",
                "collection_account": "75000",
            },
            {
                "name": "ANTIGUOS",
                "fringe": "30",
                "diary_goal": "71",
                "days": "25",
                "month_goal": "1681",
                "hours": "9",
                "collection_account": "75001",
            },
            {
                "name": "ANTIGUOS",
                "fringe": "60",
                "diary_goal": "72",
                "days": "26",
                "month_goal": "1682",
                "hours": "10",
                "collection_account": "75002",
            },
            {
                "name": "ANTIGUOS",
                "fringe": "90",
                "diary_goal": "73",
                "days": "27",
                "month_goal": "1683",
                "hours": "11",
                "collection_account": "75003",
            },
            {
                "name": "ANTIGUOS",
                "fringe": "120A180",
                "diary_goal": "74",
                "days": "28",
                "month_goal": "1684",
                "hours": "12",
                "collection_account": "75004",
            },
            {
                "name": "ANTIGUOS",
                "fringe": "210",
                "diary_goal": "75",
                "days": "29",
                "month_goal": "1685",
                "hours": "13",
                "collection_account": "75005",
            },
            {
                "name": "ANTIGUOS",
                "fringe": "PREPOTENCIAL",
                "diary_goal": "76",
                "days": "30",
                "month_goal": "1686",
                "hours": "14",
                "collection_account": "75006",
            },
            {
                "name": "ANTIGUOS",
                "fringe": "PREPOTENCIAL_2",
                "diary_goal": "77",
                "days": "31",
                "month_goal": "1687",
                "hours": "15",
                "collection_account": "75007",
            },
            {
                "name": "ANTIGUOS",
                "fringe": "PREPROVISION",
                "diary_goal": "78",
                "days": "32",
                "month_goal": "1688",
                "hours": "16",
                "collection_account": "75008",
            },
            {
                "name": "ANTIGUOS",
                "fringe": "CHURN",
                "diary_goal": "79",
                "days": "33",
                "month_goal": "1689",
                "hours": "17",
                "collection_account": "75009",
            },
            {
                "name": "ANTIGUOS",
                "fringe": "PRECHURN",
                "diary_goal": "80",
                "days": "34",
                "month_goal": "1690",
                "hours": "18",
                "collection_account": "75010",
            },
            {
                "name": "ANTIGUOS",
                "fringe": "PROVISION",
                "diary_goal": "81",
                "days": "35",
                "month_goal": "1691",
                "hours": "19",
                "collection_account": "75011",
            },
            {
                "name": "ANTIGUOS",
                "fringe": "POTENCIAL",
                "diary_goal": "82",
                "days": "36",
                "month_goal": "1692",
                "hours": "20",
                "collection_account": "75012",
            },
        ]
        # Create TableInfo objects using the data
        for data in table_info_data:
            TableInfo.objects.create(**data)
        goal = Goals.objects.create(
            cedula=settings.TEST_CEDULA,
            name="Heibert",
            campaign_goal="Base Test Goal",
            result="50",
            evaluation="50",
            quality="50",
            clean_desk="50",
            total="50",
            job_title_goal="Developer",
            last_update=timezone.now(),
            criteria_goal="50",
            quantity_goal="50",
            goal_date="ENERO-2022",
            execution_date="FEBRERO-2022",
            table_goal="ANTIGUOS",
        )
        # Prepare the request data
        payload = {
            "accepted": True,
        }
        # Send a PATCH request to the update-goal view
        response = self.client.patch(
            reverse("goal-detail", args=[goal.cedula]), data=payload
        )
        # Assert the response status code and content
        self.assertEqual(response.status_code, status.HTTP_200_OK, response.data)
        # Assert that the goal was updated with the new data
        self.assertEqual(
            response.data["message"], "La meta fue aceptada.", response.data
        )

    def test_patch_goal_execution(self):
        """Test the update-goal view."""
        # Get the first goal from the database
        goal = Goals.objects.create(
            cedula=settings.TEST_CEDULA,
            name="Heibert",
            campaign_goal="Base Test Goal",
            result="50",
            evaluation="50",
            quality="50",
            clean_desk="50",
            total="50",
            job_title_goal="Developer",
            last_update=timezone.now(),
            criteria_goal="50",
            quantity_goal="50",
            goal_date="ENERO-2022",
            execution_date="FEBRERO-2022",
        )
        # Prepare the request data
        payload = {
            "accepted_execution": True,
        }
        # Send a PATCH request to the update-goal view
        response = self.client.patch(
            reverse("goal-detail", args=[goal.cedula]), data=payload
        )
        # Assert the response status code and content
        self.assertEqual(response.status_code, status.HTTP_200_OK, response.data)
        # Assert that the goal was updated with the new data
        self.assertEqual(
            response.data["message"], "La ejecución fue aceptada.", response.data
        )

    def test_patch_goal_denied(self):
        """Test the update-goal view."""
        # Get the first goal from the database
        goal = Goals.objects.create(
            cedula=settings.TEST_CEDULA,
            name="Heibert",
            campaign_goal="Base Test Goal",
            result="50",
            evaluation="50",
            quality="50",
            clean_desk="50",
            total="50",
            job_title_goal="Developer",
            last_update=timezone.now(),
            criteria_goal="50",
            quantity_goal="50",
            goal_date="ENERO-2022",
            execution_date="FEBRERO-2022",
        )
        # Prepare the request data
        payload = {
            "result": "100",
            "evaluation": "100",
            "quality": "100",
            "clean_desk": "100",
            "total": "100",
            "job_title_goal": "Developer",
            "criteria_goal": "100",
            "quantity_goal": "100",
            "accepted": True,
        }
        # Send a PATCH request to the update-goal view
        response = self.client.patch(
            reverse("goal-detail", args=[goal.cedula]), data=payload
        )
        # Assert the response status code and content
        self.assertEqual(
            response.status_code, status.HTTP_400_BAD_REQUEST, response.data
        )
        # Assert that the goal was updated with the new data
        self.assertEqual(
            response.data["message"],
            "Patch request solo acepta el campo 'accepted' o 'accepted_execution'.",
            response.data,
        )
