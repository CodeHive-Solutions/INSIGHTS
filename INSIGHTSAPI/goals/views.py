"""This view allow to upload an Excel file with the goals of the staff and save them in the db."""

import logging
import re
import locale
from django.core.exceptions import ValidationError
from django.db import transaction
from django.db.models import Q, Subquery, Max
from django.utils import timezone
from openpyxl import load_workbook
from rest_framework import status as framework_status
from rest_framework import viewsets
from rest_framework.response import Response
from django.core.mail import send_mail
from django.core.mail import get_connection

from services.permissions import CustomizableGetDjangoModelViewPermissions


from .models import Goals, TableInfo, HistoricalGoals
from .serializers import GoalSerializer

logger = logging.getLogger("requests")
locale.setlocale(locale.LC_TIME, "es_ES.UTF-8")


class GoalsViewSet(viewsets.ModelViewSet):
    """This viewset automatically provides:
    `list`, `create`, `retrieve`, `update` and `destroy` actions.
    Additionally we also provide an extra `send_email` action."""

    queryset = Goals.objects.all()
    serializer_class = GoalSerializer
    CustomizableGetDjangoModelViewPermissions.perms_map = {
        "GET": [],
        "PATCH": [],
    }
    permission_classes = [CustomizableGetDjangoModelViewPermissions]

    def partial_update(self, request, *args, **kwargs):
        """If the user accept her goal then send him a email with the PDF file attached."""
        instance = self.get_object()
        user = request.user
        if request.data.get("accepted") is not None and len(request.data) == 1:
            if instance.accepted:
                return Response(
                    {"message": "La meta ya fue aceptada."},
                    status=framework_status.HTTP_400_BAD_REQUEST,
                )
            else:
                month = instance.goal_date.replace("-", " ").lower()
                instance.accepted_at = timezone.now()
                instance.accepted = request.data["accepted"]
                instance.save()
                table_info = TableInfo.objects.filter(name=instance.table_goal)
                accepted_state = "aceptada" if request.data["accepted"] else "rechazada"
                table_data = ""
                table_data_plain = ""
                for table in table_info:
                    table_data += f"""
                    <tr>
                        <td>{table.fringe}</td>
                        <td>{table.diary_goal}</td>
                        <td>{table.days}</td>
                        <td>{table.month_goal}</td>
                        <td>{table.hours}</td>
                        <td>{table.collection_account}</td>
                    </tr>
                    """
                    table_data_plain += f"{table.fringe:<15} | {table.diary_goal:<10} | {table.days:<4} | {table.month_goal:<18} | {table.hours:<7} | {table.collection_account}\n"
                if "CLARO" in instance.campaign_goal.upper():
                    backend = get_connection()
                    send_mail(
                        f"Meta {month}",
                        f"""
                        La meta fue {accepted_state}.

                        Información de la meta:

                        - Cedula: {instance['cedula']}
                        - Nombres: {instance['name']}
                        - Campaña: {instance['campaign_goal']}
                        - Cargo: {instance['job_title_goal']}
                        - Coordinador: {instance['coordinator_goal']}
                        - Mes: {instance['goal_date']}

                        Franja           | Meta Diaria | Días | Meta Mes con Pago | Por Hora | Recaudo por Cuenta
                        -----------------------------------------------------------------------------------------
                        {table_data_plain}
                        """,
                        None,
                        [user.email],
                        html_message=f"""
                        <p style="text-align: start">
                        La meta fue <b>{accepted_state}</b>.<br>
                        
                        Información de la meta:<br>
                        </p>
                        <ul style="padding-bottom: 1rem; text-align: start">
                            <li>Cedula: {instance.cedula}</li>
                            <li>Nombres: {instance.name}</li>
                            <li>Campaña: {instance.campaign_goal}</li>
                            <li>Cargo: {instance.job_title_goal}</li>
                            <li>Coordinador: {instance.coordinator_goal}</li>
                            <li>Mes: {instance.goal_date}</li>
                        </ul>
                        <table style="width:100%; text-align: center;">
                            <tr>
                                <th>Franja</th>
                                <th>Meta Diaria</th>
                                <th>Días</th>
                                <th>Meta Mes con Pago</th>
                                <th>Por Hora</th>
                                <th>Recaudo por Cuenta</th>
                            </tr>
                            <tr>
                                {table_data}
                            </tr>
                        </table>
                        """,
                        connection=backend,
                    )
                    print("correo enviado")
                    return Response(
                        {"message": f"La meta fue {accepted_state}."},
                        status=framework_status.HTTP_200_OK,
                    )
                else:
                    send_mail(
                        f"Meta {month}",
                        f"""
                        La meta fue {accepted_state}.

                        Información de la meta:

                        - Cedula: {instance.cedula}
                        - Nombres: {instance.name}
                        - Campaña: {instance.campaign_goal}
                        - Cargo: {instance.job_title_goal}
                        - Coordinador: {instance.coordinator_goal}
                        - Mes: {instance.goal_date}

                        Descripción de la Variable a medir | Cantidad
                        ---------------------------------- | --------
                        {instance.criteria_goal:<33} | {instance.quantity_goal}
                        """,
                        None,
                        [user.email],
                        html_message=f"""
                        <p style="text-align: start">La meta fue <b>{accepted_state}</b>.<br>
                        Información de la meta:<br>
                        </p>
                        <ul style="padding-bottom: 1rem; text-align: start">
                            <li>Cedula: {instance.cedula}</li>
                            <li>Nombres: {instance.name}</li>
                            <li>Campaña: {instance.campaign_goal}</li>
                            <li>Cargo: {instance.job_title_goal}</li>
                            <li>Coordinador: {instance.coordinator_goal}</li>
                            <li>Mes: {instance.goal_date}</li>
                        </ul>
                        <table style="width:100%; text-align: center;">
                            <tr>
                                <th>Descripción de la Variable a medir</th>
                                <th>Cantidad</th>
                            </tr>
                            <tr>
                                <td>{instance.criteria_goal}</td>
                                <td>{instance.quantity_goal}</td>
                            </tr>
                        </table>
                        """,
                    )
                    return Response(
                        {"message": "La meta fue aceptada."},
                        status=framework_status.HTTP_200_OK,
                    )
        elif (
            request.data.get("accepted_execution") is not None
            and len(request.data) == 1
        ):
            accepted_state = (
                "aceptada" if request.data["accepted_execution"] else "rechazada"
            )
            if instance.accepted_execution:
                return Response(
                    {"message": "La ejecución ya fue aceptada."},
                    status=framework_status.HTTP_400_BAD_REQUEST,
                )
            else:
                month = instance.execution_date.replace("-", " ").lower()
                instance.accepted_execution_at = timezone.now()
                instance.accepted_execution = request.data["accepted_execution"]
                instance.save()
                send_mail(
                    f"Ejecución de meta {month}",
                    f"""
                    La ejecución de la meta fue {accepted_state}.
                    
                    Información de la ejecución de la meta:
                    
                    - Cedula: {instance.cedula}
                    - Nombres: {instance.name}
                    - Campaña: {instance.campaign_execution}
                    - Cargo: {instance.job_title_execution}
                    - Coordinador: {instance.coordinator_execution}
                    - Mes: {instance.execution_date}
                    
                    Clean Desk    | Evaluación | Resultado | Calidad | Total
                    ------------- | ---------- | --------- | ------- | -----
                    {instance.clean_desk:<13} | {instance.evaluation:<10} | {instance.result:<9} | {instance.quality:<7} | {instance.total}
                    """,
                    None,
                    [user.email],
                    html_message=f"""
                    <p style="text-align: start">La ejecución de la meta fue <b>{accepted_state}</b>.<br>
                    Información de la ejecución de la meta:<br>
                    <p/>
                    <ul style="padding-bottom: 1rem; text-align: start">
                        <li>Cedula: {instance.cedula}</li>
                        <li>Nombres: {instance.name}</li>
                        <li>Campaña: {instance.campaign_execution}</li>
                        <li>Cargo: {instance.job_title_execution}</li>
                        <li>Coordinador: {instance.coordinator_execution}</li>
                        <li>Mes: {instance.execution_date}</li>
                    </ul>
                    <table style="width:100%; border-collapse: collapse; text-align: center;">
                        <tr>
                            <th>Clean Desk</th>
                            <th>Evaluación</th>
                            <th>Resultado</th>
                            <th>Calidad</th>
                            <th>Total</th>
                        </tr>
                        <tr>
                            <td>{instance.clean_desk}</td>
                            <td>{instance.evaluation}</td>
                            <td>{instance.result}</td>
                            <td>{instance.quality}</td>
                            <td>{instance.total}</td>
                        </tr>
                    </table>
                    """,
                )
                return Response(
                    {"message": f"La ejecución fue {accepted_state}."},
                    status=framework_status.HTTP_200_OK,
                )
        else:
            return Response(
                {
                    "message": "Patch request solo acepta el campo 'accepted' o 'accepted_execution'."
                },
                status=framework_status.HTTP_400_BAD_REQUEST,
            )

    def update(self, request, *args, **kwargs):
        return Response(
            {"message": "No se permite actualizar registros."},
            status=framework_status.HTTP_405_METHOD_NOT_ALLOWED,
        )

    def retrieve(self, request, *args, **kwargs):
        cedula = self.kwargs.get("pk")
        date = self.request.GET.get("date", None)
        column = self.request.GET.get("column", None)
        if request.user.cedula == cedula:
            return super().retrieve(request, *args, **kwargs)
        elif (
            date is not None
            and column is not None
            and (
                request.user.has_perm("goals.view_goals")
                or request.user.has_perm("goals.view_historicalgoals")
            )
        ):
            if column == "delivery":
                column_name = "goal_date"
            elif column == "execution":
                column_name = "execution_date"
            else:
                return self.queryset.none()
            filter_params = {f"{column_name}": date.upper()}
            # Return a single object if cedula is provided
            unique_goal = (
                HistoricalGoals.objects.filter(Q(cedula=cedula, **filter_params))
                .order_by("-history_date")
                .first()
            )
            if unique_goal:
                serializer = self.get_serializer(unique_goal)
                return Response(serializer.data)
            else:
                return Response([])  # No matching record
        elif request.user.has_perm("goals.view_goals"):
            return super().retrieve(request, *args, **kwargs)
        else:
            return Response(
                {"message": "No tiene permisos para realizar esta acción."},
                status=framework_status.HTTP_403_FORBIDDEN,
            )

    def get_queryset(self):
        if not self.request.user.has_perm("goals.view_goals"):
            return self.queryset.filter(cedula=self.request.user.cedula)
        coordinator = self.request.GET.get("coordinator", None)
        date = self.request.GET.get("date", None)
        column = self.request.GET.get("column", None)
        if coordinator is not None:
            return self.queryset.filter(coordinator=coordinator)
        elif date is not None and column is not None:
            if column == "delivery":
                column_name = "goal_date"
            elif column == "execution":
                column_name = "execution_date"
            else:
                return self.queryset.none()
            filter_params = {f"{column_name}": date.upper()}
            latest_history_dates = (
                HistoricalGoals.objects.filter(Q(**filter_params))
                .values("cedula")
                .annotate(max_history_date=Max("history_date"))
            )
            # Filter records with the latest history_date
            history_goals = HistoricalGoals.objects.filter(
                Q(
                    history_date__in=Subquery(
                        latest_history_dates.values("max_history_date")
                    )
                )
                & Q(cedula__in=latest_history_dates.values("cedula"))
            )
            return history_goals
        else:
            # With out the .all() method, the queryset will be evaluated lazily
            return self.queryset.all()

    def create(self, request, *args, **kwargs):
        # Get the file from the request
        file_obj = request.FILES.get("file")
        if file_obj:
            file_name = str(file_obj.name)
            year_pattern = r"\b\d{4}\b"
            months = [
                "ENERO",
                "FEBRERO",
                "MARZO",
                "ABRIL",
                "MAYO",
                "JUNIO",
                "JULIO",
                "AGOSTO",
                "SEPTIEMBRE",
                "OCTUBRE",
                "NOVIEMBRE",
                "DICIEMBRE",
            ]
            month_pattern = r"(?i)(" + "|".join(months) + ")"
            if not re.search(month_pattern, file_name):
                return Response(
                    {"message": "Mes no encontrado en el nombre del archivo."},
                    status=framework_status.HTTP_400_BAD_REQUEST,
                )
            elif not re.search(year_pattern, file_name):
                return Response(
                    {"message": "Año no encontrado en el nombre del archivo."},
                    status=framework_status.HTTP_400_BAD_REQUEST,
                )
            try:
                date = (
                    file_name.split("-")[1].upper()
                    + "-"
                    + file_name.split("-")[2].split(".")[0]
                )
                # Read the Excel file using openpyxl
                workbook = load_workbook(file_obj, read_only=True, data_only=True)
                sheet = workbook[workbook.sheetnames[0]]
                # Get the column indices based on the header names
                header_row = next(sheet.iter_rows(values_only=True))
                header_names = {
                    "CEDULA",
                    "NOMBRE",
                    "CARGO",
                    "CAMPAÑA",
                    "COORDINADOR A CARGO",
                    "ESTADO",
                    "OBSERVACIONES",
                }
                missing_headers = header_names - set(header_row)  # type: ignore
                if missing_headers:
                    return Response(
                        {
                            "message": f"Columnas no encontradas: {', '.join(missing_headers)}"
                        },
                        status=framework_status.HTTP_400_BAD_REQUEST,
                    )
                cedula_index = header_row.index("CEDULA")
                name_index = header_row.index("NOMBRE")
                cargo_index = header_row.index("CARGO")
                campaign_index = header_row.index("CAMPAÑA")
                coordinator_index = header_row.index("COORDINADOR A CARGO")
                status_index = header_row.index("ESTADO")
                observation_index = header_row.index("OBSERVACIONES")
                if file_name.upper().find("CLARO") != -1:
                    with transaction.atomic():
                        header_names = {"TABLA"}
                        header_row = next(sheet.iter_rows(values_only=True))  # type: ignore
                        missing_headers = header_names - set(header_row)  # type: ignore
                        if missing_headers:
                            return Response(
                                {
                                    "message": f"Encabezados no encontrados: {', '.join(missing_headers)}"
                                },
                                status=framework_status.HTTP_400_BAD_REQUEST,
                            )
                        table_index = header_row.index("TABLA")
                        sheet = workbook["Tabla_de_valores"]
                        header_names = {
                            "NOMBRE_TABLA",
                            "FRANJA",
                            "META DIARIA",
                            "DIAS",
                            "META MES CON PAGO",
                            "POR HORA",
                            "RECAUDO POR CUENTA",
                        }
                        header_row = next(sheet.iter_rows(values_only=True))  # type: ignore
                        missing_headers = header_names - set(header_row)  # type: ignore
                        if missing_headers:
                            return Response(
                                {
                                    "message": f"Encabezados no encontrados: {', '.join(missing_headers)}"
                                },
                                status=framework_status.HTTP_400_BAD_REQUEST,
                            )
                        table_name_index = header_row.index("NOMBRE_TABLA")
                        fringe_index = header_row.index("FRANJA")
                        diary_goal_index = header_row.index("META DIARIA")
                        days_index = header_row.index("DIAS")
                        month_goal_index = header_row.index("META MES CON PAGO")
                        hours_index = header_row.index("POR HORA")
                        collection_account_index = header_row.index(
                            "RECAUDO POR CUENTA"
                        )
                        sheet = workbook[workbook.sheetnames[0]]
                        for i, row in enumerate(sheet.iter_rows(min_row=2), start=2):  # type: ignore
                            cargo_goal = str(row[cargo_index].value).upper().lstrip(".")
                            if cargo_goal.find("ASESOR") != -1:
                                # Avoid NoneType error
                                def format_cell_value(cell):
                                    if cell.value is None or "":
                                        return ""
                                    else:
                                        return "{:.2%}".format(cell.value)

                                cedula = row[cedula_index].value
                                name = row[name_index].value
                                coordinator_goal = row[coordinator_index].value
                                campaign_goal = row[campaign_index].value
                                observation_goal = row[observation_index].value
                                table = row[table_index].value
                                status_goal = row[status_index].value
                                if str(status_goal).upper() == "ACTIVO":
                                    status_goal = True
                                elif str(status_goal).upper() == "RETIRADO":
                                    status_goal = False
                                else:
                                    return Response(
                                        {
                                            "message": "Se encontraron asesores sin estado."
                                        },
                                        status=framework_status.HTTP_400_BAD_REQUEST,
                                    )
                                # Update or create the record
                                unique_constraint = "cedula"
                                default_value = {
                                    "name": str(name).upper(),
                                    "job_title_goal": cargo_goal,
                                    "campaign_goal": campaign_goal,
                                    "coordinator_goal": coordinator_goal,
                                    "goal_date": date,
                                    "status_goal": status_goal,
                                    "observation_goal": observation_goal,
                                    "table_goal": table,
                                }
                                try:
                                    instance = Goals.objects.filter(cedula=cedula)
                                    if instance.exists():
                                        instance.update(accepted=None, accepted_at=None)
                                    Goals.objects.update_or_create(
                                        defaults=default_value,
                                        **{unique_constraint: cedula},
                                    )
                                except ValidationError as validation_e:
                                    logger.setLevel(logging.ERROR)
                                    logger.exception(
                                        "Validation error: %s", str(validation_e)
                                    )
                                    return Response(
                                        {
                                            "message": "Excel upload Failed.",
                                        },
                                        status=framework_status.HTTP_400_BAD_REQUEST,
                                    )
                                except Exception as error:
                                    logger.setLevel(logging.ERROR)
                                    logger.exception("Error: %s", str(error))
                                    return Response(
                                        {
                                            "message": "Excel upload Failed.",
                                        },
                                        status=framework_status.HTTP_500_INTERNAL_SERVER_ERROR,
                                    )
                        sheet = workbook["Tabla_de_valores"]
                        table_name = None
                        for i, row in enumerate(sheet.iter_rows(min_row=2), start=2):  # type: ignore
                            if i == 2:
                                TableInfo.objects.all().delete()
                            try:
                                if row[table_name_index].value is not None or "":
                                    table_name = row[table_name_index].value
                                if table_name:
                                    fringe = row[fringe_index].value
                                    diary_goal = row[diary_goal_index].value
                                    days = row[days_index].value
                                    month_goal = row[month_goal_index].value
                                    hours = row[hours_index].value
                                    collection_account_str = str(
                                        row[collection_account_index].value
                                    )
                                    collection_account = re.sub(
                                        r"[^0-9]", "", collection_account_str
                                    )
                                    TableInfo.objects.create(
                                        name=str(table_name).upper(),
                                        fringe=fringe,
                                        diary_goal=diary_goal,
                                        days=days,
                                        month_goal=month_goal,
                                        hours=hours,
                                        collection_account=collection_account,
                                    )
                            except Exception as error:
                                logger.setLevel(logging.ERROR)
                                logger.exception("Error: %s", str(error))
                                return Response(
                                    {"message": "Excel upload Failed."},
                                    status=framework_status.HTTP_500_INTERNAL_SERVER_ERROR,
                                )
                        return Response(
                            {
                                "message": "Excel file uploaded and processed successfully."
                            },
                            status=framework_status.HTTP_201_CREATED,
                        )
                header_names = {"DESCRIPCION DE LA VARIABLE A MEDIR", "CANTIDAD"}
                missing_headers = header_names - set(header_row)  # type: ignore
                if missing_headers:
                    return Response(
                        {
                            "message": f"Estos encabezados no fueron encontrados: {', '.join(missing_headers)}"
                        },
                        status=framework_status.HTTP_400_BAD_REQUEST,
                    )
                criteria_index = header_row.index("DESCRIPCION DE LA VARIABLE A MEDIR")
                quantity_index = header_row.index("CANTIDAD")
                if (
                    file_name.upper().find("EJECUCIÓN") != -1
                    or file_name.upper().find("EJECUCION") != -1
                ):
                    header_names = {
                        "% CUMPLIMIENTO",
                        "EVALUACION",
                        "CALIDAD",
                        "CLEAN DESK",
                        "TOTAL",
                    }
                    missing_headers = header_names - set(header_row)  # type: ignore
                    if missing_headers:
                        return Response(
                            {
                                "message": f"Estos encabezados no fueron encontrados: {', '.join(missing_headers)}"
                            },
                            status=framework_status.HTTP_400_BAD_REQUEST,
                        )
                    result_index = header_row.index("% CUMPLIMIENTO")
                    evaluation_index = header_row.index("EVALUACION")
                    quality_index = header_row.index("CALIDAD")
                    clean_desk_index = header_row.index("CLEAN DESK")
                    total_index = header_row.index("TOTAL")
                else:
                    result_index = None
                    evaluation_index = None
                    quality_index = None
                    clean_desk_index = None
                    total_index = None
                for i, row in enumerate(sheet.iter_rows(min_row=2), start=2):  # type: ignore
                    cargo = str(row[cargo_index].value)
                    cedula = row[cedula_index]
                    if cargo.upper().find("ASESOR") != -1:
                        # Avoid NoneType error
                        def format_cell_value(cell):
                            if cell.value is None or "":
                                return ""
                            else:
                                return "{:.2%}".format(cell.value)

                        cedula = row[cedula_index].value
                        name = row[name_index].value

                        entrega = True
                        unique_constraint = "cedula"
                        if (
                            file_name.upper().find("EJECUCION") != -1
                            or file_name.upper().find("EJECUCIÓN") != -1
                        ):
                            cargo_execution = cargo.upper()
                            coordinator_execution = row[coordinator_index].value
                            campaign_execution = row[campaign_index].value
                            criteria_execution = row[criteria_index].value
                            observation_execution = row[observation_index].value
                            status_execution = row[status_index].value
                            result_cell = row[result_index]  # type: ignore
                            result = format_cell_value(result_cell)
                            evaluation_cell = row[evaluation_index]  # type: ignore
                            evaluation = format_cell_value(evaluation_cell)
                            quality = format_cell_value(row[quality_index])  # type: ignore
                            clean_desk = format_cell_value(row[clean_desk_index])  # type: ignore
                            total = format_cell_value(row[total_index])  # type: ignore
                            entrega = False
                            if str(status_execution).upper() == "ACTIVO":
                                status_execution = True
                            elif str(status_execution).upper() == "RETIRADO":
                                status_execution = False
                            else:
                                return Response(
                                    {"message": "Se encontraron asesores sin estado."},
                                    status=framework_status.HTTP_400_BAD_REQUEST,
                                )
                            default_value = {
                                "name": name,
                                "observation_execution": observation_execution,
                                "job_title_execution": cargo_execution,
                                "campaign_execution": campaign_execution,
                                "coordinator_execution": coordinator_execution,
                                "criteria_execution": criteria_execution,
                                "result": result,
                                "quality": quality,
                                "evaluation": evaluation,
                                "clean_desk": clean_desk,
                                "total": total,
                                "quantity_execution": row[quantity_index].value,
                                "status_execution": status_execution,
                            }
                        else:
                            cargo_goal = cargo.upper()
                            coordinator_goal = row[coordinator_index].value
                            campaign_goal = row[campaign_index].value
                            criteria_goal = row[criteria_index].value
                            observation_goal = row[observation_index].value
                            status_goal = row[status_index].value
                            result = None
                            evaluation = None
                            quality = None
                            clean_desk = None
                            total = None
                            if str(status_goal).upper() == "ACTIVO":
                                status_goal = True
                            elif str(status_goal).upper() == "RETIRADO":
                                status_goal = False
                            else:
                                return Response(
                                    {"message": "Se encontraron asesores sin estado."},
                                    status=framework_status.HTTP_400_BAD_REQUEST,
                                )
                            default_value = {
                                "name": name,
                                "observation_goal": observation_goal,
                                "job_title_goal": cargo_goal,
                                "campaign_goal": campaign_goal,
                                "coordinator_goal": coordinator_goal,
                                "criteria_goal": criteria_goal,
                                "result": result,
                                "quality": quality,
                                "evaluation": evaluation,
                                "clean_desk": clean_desk,
                                "total": total,
                                "quantity_goal": row[quantity_index].value,
                                "status_goal": status_goal,
                            }
                        # Remove empty values from default_value dictionary
                        default_value = {k: v for k, v in default_value.items() if v}
                        instance = Goals.objects.filter(cedula=cedula)
                        if file_name.upper().find("ENTREGA") != -1:
                            default_value["goal_date"] = date
                            if instance.exists():
                                instance.update(accepted=None, accepted_at=None)
                        elif (
                            file_name.upper().find("EJECUCION") != -1
                            or file_name.upper().find("EJECUCIÓN") != -1
                        ):
                            default_value["execution_date"] = date
                            if instance.exists():
                                instance.update(
                                    accepted_execution=None, accepted_execution_at=None
                                )
                        Goals.objects.update_or_create(
                            defaults=default_value, **{unique_constraint: cedula}
                        )
                return Response(
                    {"message": "Excel file uploaded and processed successfully."},
                    status=framework_status.HTTP_201_CREATED,
                )
            except ValidationError as validation_e:
                logger.setLevel(logging.ERROR)
                logger.exception("Validation error: %s", str(validation_e))
                return Response(
                    {"message": "Excel upload Failed."},
                    status=framework_status.HTTP_400_BAD_REQUEST,
                )
            except Exception as error:
                logger.setLevel(logging.ERROR)
                logger.exception("Error: %s", str(error))
                return Response(
                    {"message": "Excel upload Failed."},
                    status=framework_status.HTTP_500_INTERNAL_SERVER_ERROR,
                )
        else:
            return Response(
                {"message": "Excel no encontrado."},
                status=framework_status.HTTP_400_BAD_REQUEST,
            )
